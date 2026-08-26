"""Build the TCS five-year scenario workbook — fully formula-driven."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from model import (BUCKETS, DRIVERS, PHASING, MARGINS, FCF_CONV, PAYOUT_OF_FCF,
                   REV_PER_EMP_GROWTH, PROBABILITY, YEARS, HISTORY, FX_FY26,
                   FX_DEPRECIATION, OTHER_INCOME_MARGIN, TAX_RATE, TATA_SONS_STAKE,
                   SHARES_CR, REV_PER_EMP_FY26, TS_NON_TCS_DIV_FY26, TS_NON_TCS_GROWTH,
                   TS_OPEX_TAX_FY26, TS_OPEX_GROWTH, TS_NET_CASH_FY26, TS_CAPITAL_CALL,
                   GROWTH_DECAY)

FONT = "Arial"
BLUE = Font(name=FONT, size=10, color="0000FF")
BLACK = Font(name=FONT, size=10)
GREEN = Font(name=FONT, size=10, color="008000")
BOLD = Font(name=FONT, size=10, bold=True)
TITLE = Font(name=FONT, size=14, bold=True, color="1F3864")
H1 = Font(name=FONT, size=11, bold=True, color="FFFFFF")
NOTE = Font(name=FONT, size=9, italic=True, color="595959")
HDRFILL = PatternFill("solid", fgColor="1F3864")
SUBFILL = PatternFill("solid", fgColor="D9E2F3")
YELLOW = PatternFill("solid", fgColor="FFFF00")
GREYFILL = PatternFill("solid", fgColor="F2F2F2")
THIN = Side(style="thin", color="BFBFBF")
BOX = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)

USD = '$#,##0;($#,##0);"-"'
USD1 = '$#,##0.0;($#,##0.0);"-"'
INR = '#,##0;(#,##0);"-"'
PCT = '0.0%;(0.0%);"-"'
PCT2 = '0.00%;(0.00%);"-"'
NUM = '#,##0;(#,##0);"-"'
RUPEE = '"₹"#,##0.0;("₹"#,##0.0);"-"'

SCENS = list(DRIVERS.keys())
SC_SHORT = {"Bull — AI Dividend": "Bull", "Base — Managed Transition": "Base",
            "Bear — Slow Bleed": "Bear", "Severe — Structural Break": "Severe"}

wb = openpyxl.Workbook()


def style_all(ws):
    for row in ws.iter_rows():
        for c in row:
            if c.font is None or c.font.name != FONT:
                if c.font and c.font.color and c.font.color.rgb in ("FF0000FF", "FF008000"):
                    continue
                c.font = Font(name=FONT, size=10)


def header(ws, row, text, width):
    ws.cell(row, 1, text).font = H1
    for c in range(1, width + 1):
        ws.cell(row, c).fill = HDRFILL
        ws.cell(row, c).font = H1


# =============================================================================
# READ ME
# =============================================================================
ws = wb.active
ws.title = "Read Me"
ws.column_dimensions["A"].width = 3
ws.column_dimensions["B"].width = 34
ws.column_dimensions["C"].width = 96
ws["B2"] = "TCS — Five-Year Scenario Model (FY27–FY31)"
ws["B2"].font = TITLE
ws["B3"] = "Prepared for the Chairman, Tata Sons  ·  base year FY26 (year ended 31 March 2026)"
ws["B3"].font = NOTE

rows = [
    ("", ""),
    ("WHAT THIS IS", "A driver-based model of TCS revenue, margin, cash and dividend under four scenarios, "
                     "and what each means for Tata Sons' parent-level funding capacity."),
    ("HOW TO USE IT", "Every blue cell is an input you can change. Everything black is a formula. "
                      "Change an assumption on 'Assumptions' and every sheet recalculates."),
    ("", ""),
    ("COLOUR LEGEND", ""),
    ("  Blue text", "Hard-coded input / scenario lever — edit these."),
    ("  Black text", "Formula — do not overwrite."),
    ("  Green text", "Link to another sheet in this workbook."),
    ("  Yellow fill", "Key judgement call. These drive the answer; challenge them first."),
    ("", ""),
    ("SHEETS", ""),
    ("  History", "TCS actuals FY22–FY26 and Q1 FY27, with sources."),
    ("  Assumptions", "All shared parameters, the revenue decomposition, and the four scenario driver blocks."),
    ("  Revenue Build", "Bucket-level revenue build for each scenario, FY27–FY31."),
    ("  Scenario Output", "Full P&L, cash, dividend and headcount output for each scenario."),
    ("  Tata Sons Bridge", "Parent-level cash bridge: TCS dividend vs. the capital call from Air India, "
                           "Tata Electronics, Agratas and Tata Digital."),
    ("  Sensitivity", "FY31 dividend to Tata Sons across revenue CAGR × terminal operating margin."),
    ("  Risk Register", "Revenue-at-risk by exposure bucket."),
    ("", ""),
    ("THE ONE CAVEAT", "TCS stopped disclosing a service-line revenue split after FY21. The seven-bucket "
                       "decomposition on 'Assumptions' is a reconstruction, not a disclosure. It is the single "
                       "most challengeable input in this model and the first thing to replace with internal MIS."),
    ("", ""),
    ("PRIMARY SOURCES", "TCS Q4 FY26 Fact Sheet and press release; TCS Q1 FY27 press release; "
                        "Tata Sons FY26 accounts as reported; Business Standard; Business Today; company filings."),
]
r = 5
for a, b in rows:
    ws.cell(r, 2, a).font = BOLD if a and not a.startswith("  ") else BLACK
    ws.cell(r, 3, b).font = BLACK
    ws.cell(r, 3).alignment = Alignment(wrap_text=True, vertical="top")
    if a.strip() in ("Blue text",):
        ws.cell(r, 2).font = BLUE
    if a.strip() in ("Green text",):
        ws.cell(r, 2).font = GREEN
    if a.strip() in ("Yellow fill",):
        ws.cell(r, 2).fill = YELLOW
    ws.row_dimensions[r].height = 28 if len(b) > 90 else 15
    r += 1

# =============================================================================
# HISTORY
# =============================================================================
ws = wb.create_sheet("History")
ws.column_dimensions["A"].width = 42
for i in range(2, 9):
    ws.column_dimensions[get_column_letter(i)].width = 13
ws["A1"] = "TCS — Reported actuals"
ws["A1"].font = TITLE
ws["A2"] = "Source: TCS quarterly fact sheets and results press releases, tcs.com/investor-relations"
ws["A2"].font = NOTE

hdr = ["Metric"] + list(HISTORY.FY) + ["Q1 FY27"]
header(ws, 4, "", len(hdr))
for j, h in enumerate(hdr, 1):
    ws.cell(4, j, h)
    ws.cell(4, j).font = H1
    ws.cell(4, j).alignment = Alignment(horizontal="center")

hist_rows = [
    ("Revenue ($mn)", "RevUSDm", USD, 7624),
    ("Revenue growth, reported USD", "USDg", PCT, None),
    ("Revenue growth, constant currency", "CCg", PCT, 0.032),
    ("Operating margin", "OPM", PCT, 0.240),
    ("Net margin", "NetM", PCT, 0.192),
    ("Net income ($mn)", "PATusdm", USD, 1460),
    ("Revenue (₹ crore)", "RevINRcr", INR, None),
    ("Net income (₹ crore)", "PATINRcr", INR, None),
    ("Order book / TCV ($bn)", "TCVusdbn", '#,##0.0', 9.5),
    ("Closing headcount", "Headcount", NUM, 593798),
    ("Dividend per share (₹)", "DPS", RUPEE, 12),
    ("Total shareholder payout (₹ crore)", "PayoutINRcr", INR, None),
]
r = 5
for label, col, fmt, q1 in hist_rows:
    ws.cell(r, 1, label).font = BLACK
    for j, (_, row) in enumerate(HISTORY.iterrows(), 2):
        v = row[col]
        if col == "TCVusdbn":
            v = v / 1000 if v > 1000 else v
        c = ws.cell(r, j, float(v))
        c.font = BLUE
        c.number_format = fmt
    c = ws.cell(r, 7, q1 if q1 is not None else None)
    c.font = BLUE
    c.number_format = fmt
    r += 1

r += 1
ws.cell(r, 1, "Derived").font = BOLD
r += 1
ws.cell(r, 1, "Revenue per employee ($)").font = BLACK
for j in range(2, 7):
    c = ws.cell(r, j, f"={get_column_letter(j)}5*1000000/{get_column_letter(j)}14")
    c.number_format = '$#,##0'
r += 1
ws.cell(r, 1, "Implied ₹/US$ (revenue-weighted)").font = BLACK
for j in range(2, 7):
    c = ws.cell(r, j, f"={get_column_letter(j)}11*10/{get_column_letter(j)}5")
    c.number_format = '0.00'
r += 1
ws.cell(r, 1, "Payout as % of net income").font = BLACK
for j in range(2, 7):
    c = ws.cell(r, j, f"={get_column_letter(j)}16/{get_column_letter(j)}12")
    c.number_format = PCT
r += 2
ws.cell(r, 1, "Note: FY26 was the first US-dollar revenue decline in TCS's listed history "
              "(-0.5% reported, -2.4% constant currency).").font = NOTE

# =============================================================================
# ASSUMPTIONS
# =============================================================================
ws = wb.create_sheet("Assumptions")
A = "Assumptions"
ws.column_dimensions["A"].width = 46
for i in range(2, 10):
    ws.column_dimensions[get_column_letter(i)].width = 13
ws.column_dimensions["J"].width = 60
ws["A1"] = "Assumptions and scenario drivers"
ws["A1"].font = TITLE
ws["A2"] = "Blue = input. Yellow = key judgement. Change these and the whole model moves."
ws["A2"].font = NOTE

def put(r, label, value, fmt=None, note=None, key=False, col=2):
    ws.cell(r, 1, label).font = BLACK
    c = ws.cell(r, col, value)
    c.font = BLUE
    if fmt:
        c.number_format = fmt
    if key:
        c.fill = YELLOW
    if note:
        ws.cell(r, 10, note).font = NOTE
    return c

header(ws, 4, "1. Shared parameters", 9)
put(5, "FY26 revenue ($mn)  [actual]", float(HISTORY.iloc[-1].RevUSDm), USD, "TCS Q4 FY26 press release")
put(6, "FY26 closing headcount  [actual]", int(HISTORY.iloc[-1].Headcount), NUM, "TCS Q4 FY26 fact sheet")
put(7, "FY26 revenue per employee ($)", round(REV_PER_EMP_FY26, 0), '$#,##0', "Derived")
put(8, "FY26 implied ₹/US$", round(FX_FY26, 2), '0.00', "₹2,67,021 cr / $30,017mn")
put(9, "₹ depreciation p.a.", FX_DEPRECIATION, PCT, "Long-run inflation differential", key=True)
put(10, "Other income as % of revenue", OTHER_INCOME_MARGIN, PCT, "Treasury income on ~₹50,000 cr cash")
put(11, "Effective tax rate", TAX_RATE, PCT, "FY26 effective rate")
put(12, "Tata Sons stake in TCS", TATA_SONS_STAKE, '0.00%', "2,59,54,99,419 shares = 71.74%")
put(13, "TCS shares outstanding (crore)", SHARES_CR, '#,##0.0', "For DPS calculation")

header(ws, 15, "2. Revenue decomposition by AI exposure  (FY26 base)", 9)
for j, h in enumerate(["Exposure bucket", "FY26 weight", "FY26 $mn", "Automatable share",
                       "Gross rev at risk $mn", "Volume decay p.a."], 1):
    ws.cell(16, j, h).font = BOLD
    ws.cell(16, j).fill = SUBFILL
    ws.cell(16, j).alignment = Alignment(wrap_text=True, horizontal="center")
ws.row_dimensions[16].height = 30
BR0 = 17
for i, (name, w, auto, note) in enumerate(BUCKETS):
    r = BR0 + i
    ws.cell(r, 1, name).font = BLACK
    c = ws.cell(r, 2, w); c.font = BLUE; c.number_format = PCT; c.fill = YELLOW
    c = ws.cell(r, 3, f"=B{r}*$B$5"); c.number_format = USD
    c = ws.cell(r, 4, auto); c.font = BLUE; c.number_format = PCT; c.fill = YELLOW
    c = ws.cell(r, 5, f"=C{r}*D{r}"); c.number_format = USD
    c = ws.cell(r, 6, GROWTH_DECAY.get(i, 1.0)); c.font = BLUE; c.number_format = '0.00'
    ws.cell(r, 10, note).font = NOTE
    ws.cell(r, 10).alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[r].height = 30
rT = BR0 + len(BUCKETS)
ws.cell(rT, 1, "Total").font = BOLD
for col in (2, 3, 5):
    c = ws.cell(rT, col, f"=SUM({get_column_letter(col)}{BR0}:{get_column_letter(col)}{rT-1})")
    c.font = BOLD
    c.number_format = PCT if col == 2 else USD
ws.cell(rT + 1, 1, "TCS has not disclosed a service-line revenue split since FY21. Weights are a "
                   "reconstruction from the last disclosure, deal mix and peer structure — replace with internal MIS.").font = NOTE

# ---- scenario driver blocks ----
SC_ROWS = {}
r = rT + 4
for s in SCENS:
    header(ws, r, f"3. Scenario drivers — {s}", 9)
    ws.cell(r, 8, "Probability").font = H1
    ws.cell(r, 9, PROBABILITY[s]).font = Font(name=FONT, size=10, bold=True, color="FFFF00")
    ws.cell(r, 9).number_format = PCT
    ws.cell(r, 9).fill = HDRFILL
    base = r + 1
    ws.cell(base, 1, "Per-bucket annual drivers").font = BOLD
    ws.cell(base, 2, "Realisation deflation").font = BOLD
    ws.cell(base, 3, "Volume growth").font = BOLD
    for c in (2, 3):
        ws.cell(base, c).fill = SUBFILL
        ws.cell(base, c).alignment = Alignment(wrap_text=True, horizontal="center")
    ws.cell(base, 1).fill = SUBFILL
    ws.row_dimensions[base].height = 28
    for i, (name, *_rest) in enumerate(BUCKETS):
        rr = base + 1 + i
        ws.cell(rr, 1, name).font = BLACK
        d, v = DRIVERS[s][i]
        c = ws.cell(rr, 2, d); c.font = BLUE; c.number_format = PCT; c.fill = YELLOW
        c = ws.cell(rr, 3, v); c.font = BLUE; c.number_format = PCT; c.fill = YELLOW
    yr0 = base + 1 + len(BUCKETS)
    ws.cell(yr0, 1, "Year-specific parameters").font = BOLD
    ws.cell(yr0, 1).fill = SUBFILL
    for j, y in enumerate(YEARS):
        ws.cell(yr0, 2 + j, y).font = BOLD
        ws.cell(yr0, 2 + j).fill = SUBFILL
        ws.cell(yr0, 2 + j).alignment = Alignment(horizontal="center")
    spec = [("Deflation phasing multiplier", PHASING[s], '0.00'),
            ("Operating margin", MARGINS[s], PCT),
            ("Free cash flow / net income", FCF_CONV[s], PCT),
            ("Dividend as % of free cash flow", PAYOUT_OF_FCF[s], PCT)]
    for k, (lab, vals, fmt) in enumerate(spec):
        rr = yr0 + 1 + k
        ws.cell(rr, 1, lab).font = BLACK
        for j, v in enumerate(vals):
            c = ws.cell(rr, 2 + j, v); c.font = BLUE; c.number_format = fmt
            if lab == "Operating margin":
                c.fill = YELLOW
    rr = yr0 + 1 + len(spec)
    ws.cell(rr, 1, "Revenue per employee growth p.a.").font = BLACK
    c = ws.cell(rr, 2, REV_PER_EMP_GROWTH[s]); c.font = BLUE; c.number_format = PCT; c.fill = YELLOW
    SC_ROWS[s] = dict(block=r, drv0=base + 1, yr0=yr0, phase=yr0 + 1, marg=yr0 + 2,
                      fcf=yr0 + 3, pay=yr0 + 4, rpe=rr, prob=(r, 9))
    r = rr + 3

header(ws, r, "4. Tata Sons parent-level assumptions (₹ crore)", 9)
TS0 = r + 1
put(TS0, "Non-TCS dividend income, FY26 [actual]", TS_NON_TCS_DIV_FY26, INR,
    "Total FY26 dividend income ₹32,528 cr less TCS ₹28,291 cr")
put(TS0 + 1, "Non-TCS dividend growth p.a.", TS_NON_TCS_GROWTH, PCT,
    "Tata Motors, Tata Steel, Titan, Tata Power, TCPL, IHCL", key=True)
put(TS0 + 2, "Parent opex, interest and tax, FY26", TS_OPEX_TAX_FY26, INR,
    "Dividend income less reported operating cash flow of ₹25,544 cr")
put(TS0 + 3, "Parent cost growth p.a.", TS_OPEX_GROWTH, PCT)
put(TS0 + 4, "Opening net cash, FY26 [actual]", TS_NET_CASH_FY26, INR, "No borrowings at FY26 close")
ws.cell(TS0 + 6, 1, "Equity deployment into strategic bets (₹ crore)").font = BOLD
ws.cell(TS0 + 6, 1).fill = SUBFILL
for j, y in enumerate(YEARS):
    ws.cell(TS0 + 6, 2 + j, y).font = BOLD
    ws.cell(TS0 + 6, 2 + j).fill = SUBFILL
    ws.cell(TS0 + 6, 2 + j).alignment = Alignment(horizontal="center")
ws.cell(TS0 + 7, 1, "Air India, Tata Electronics, Agratas, Tata Digital").font = BLACK
for j, y in enumerate(YEARS):
    c = ws.cell(TS0 + 7, 2 + j, TS_CAPITAL_CALL[y]); c.font = BLUE; c.number_format = INR; c.fill = YELLOW
ws.cell(TS0 + 8, 1, "FY26 actual deployment was ₹15,956 cr (₹15,089 cr subsidiaries + ₹867 cr associates/JVs). "
                    "Forward path reflects the Dholera fab (₹91,000 cr), Assam ATMP (₹27,000 cr), Air India fleet "
                    "and Agratas. Held constant across scenarios by design — the commitments do not shrink because TCS slows.").font = NOTE
ws.cell(TS0 + 8, 1).alignment = Alignment(wrap_text=True, vertical="top")
ws.row_dimensions[TS0 + 8].height = 42
TS_CAP_ROW = TS0 + 7

# =============================================================================
# REVENUE BUILD
# =============================================================================
ws = wb.create_sheet("Revenue Build")
ws.column_dimensions["A"].width = 46
for i in range(2, 9):
    ws.column_dimensions[get_column_letter(i)].width = 14
ws["A1"] = "Revenue build by exposure bucket ($mn)"
ws["A1"].font = TITLE
ws["A2"] = "Each year = prior year × (1 + volume growth × decay^n) × (1 + deflation × phasing)"
ws["A2"].font = NOTE

RB_ROWS = {}
r = 4
for s in SCENS:
    header(ws, r, s, 7)
    ws.cell(r + 1, 1, "Exposure bucket").font = BOLD
    ws.cell(r + 1, 1).fill = SUBFILL
    ws.cell(r + 1, 2, "FY26").font = BOLD
    ws.cell(r + 1, 2).fill = SUBFILL
    ws.cell(r + 1, 2).alignment = Alignment(horizontal="center")
    for j, y in enumerate(YEARS):
        ws.cell(r + 1, 3 + j, y).font = BOLD
        ws.cell(r + 1, 3 + j).fill = SUBFILL
        ws.cell(r + 1, 3 + j).alignment = Alignment(horizontal="center")
    sr = SC_ROWS[s]
    b0 = r + 2
    for i, (name, *_x) in enumerate(BUCKETS):
        rr = b0 + i
        ws.cell(rr, 1, name).font = BLACK
        c = ws.cell(rr, 2, f"={A}!C{BR0+i}"); c.font = GREEN; c.number_format = USD
        for j in range(5):
            col = get_column_letter(3 + j)
            prev = get_column_letter(2 + j)
            f = (f"={prev}{rr}"
                 f"*(1+{A}!$C${sr['drv0']+i}*{A}!$F${BR0+i}^{j})"
                 f"*(1+{A}!$B${sr['drv0']+i}*{A}!{get_column_letter(2+j)}${sr['phase']})")
            c = ws.cell(rr, 3 + j, f); c.number_format = USD
    rt = b0 + len(BUCKETS)
    ws.cell(rt, 1, "Total revenue ($mn)").font = BOLD
    for j in range(6):
        col = get_column_letter(2 + j)
        c = ws.cell(rt, 2 + j, f"=SUM({col}{b0}:{col}{rt-1})")
        c.font = BOLD
        c.number_format = USD
    rg = rt + 1
    ws.cell(rg, 1, "Growth, constant currency").font = BLACK
    for j in range(5):
        col = get_column_letter(3 + j)
        prev = get_column_letter(2 + j)
        c = ws.cell(rg, 3 + j, f"={col}{rt}/{prev}{rt}-1"); c.number_format = PCT
    rc = rg + 1
    ws.cell(rc, 1, "FY26–FY31 revenue CAGR").font = BLACK
    c = ws.cell(rc, 3, f"=(G{rt}/B{rt})^(1/5)-1"); c.number_format = PCT; c.font = BOLD
    RB_ROWS[s] = dict(b0=b0, total=rt, growth=rg, cagr=rc)
    r = rc + 3

# =============================================================================
# SCENARIO OUTPUT
# =============================================================================
ws = wb.create_sheet("Scenario Output")
ws.column_dimensions["A"].width = 46
for i in range(2, 9):
    ws.column_dimensions[get_column_letter(i)].width = 14
ws["A1"] = "Scenario output — P&L, cash, dividend and headcount"
ws["A1"].font = TITLE
ws["A2"] = "FY26 column is actual. FY27–FY31 are model output."
ws["A2"].font = NOTE

SO_ROWS = {}
r = 4
for s in SCENS:
    sr, rb = SC_ROWS[s], RB_ROWS[s]
    header(ws, r, s, 7)
    ws.cell(r + 1, 1, "").fill = SUBFILL
    ws.cell(r + 1, 2, "FY26A").font = BOLD
    ws.cell(r + 1, 2).fill = SUBFILL
    ws.cell(r + 1, 2).alignment = Alignment(horizontal="center")
    for j, y in enumerate(YEARS):
        ws.cell(r + 1, 3 + j, y).font = BOLD
        ws.cell(r + 1, 3 + j).fill = SUBFILL
        ws.cell(r + 1, 3 + j).alignment = Alignment(horizontal="center")
    b = r + 2
    L = {}
    def line(off, label, fmt, bold=False):
        rr = b + off
        ws.cell(rr, 1, label).font = BOLD if bold else BLACK
        L[label] = rr
        return rr

    rRev = line(0, "Revenue ($mn)", USD, True)
    rG = line(1, "  growth, constant currency", PCT)
    rOPM = line(2, "Operating margin", PCT)
    rEBIT = line(3, "Operating income ($mn)", USD)
    rPBT = line(4, "Profit before tax ($mn)", USD)
    rPAT = line(5, "Net income ($mn)", USD, True)
    rNM = line(6, "  net margin", PCT)
    rFX = line(7, "₹ / US$", '0.00')
    rRI = line(8, "Revenue (₹ crore)", INR)
    rPI = line(9, "Net income (₹ crore)", INR)
    rFCF = line(10, "Free cash flow (₹ crore)", INR)
    rDIV = line(11, "Total dividend (₹ crore)", INR, True)
    rDPS = line(12, "Dividend per share (₹)", RUPEE)
    rTS = line(13, "Dividend to Tata Sons (₹ crore)", INR, True)
    rRPE = line(14, "Revenue per employee ($)", '$#,##0')
    rHC = line(15, "Closing headcount", NUM)
    rHCd = line(16, "  change vs FY26", NUM)

    # FY26 actuals column
    fy26 = HISTORY.iloc[-1]
    for rr, val, fmt in [(rRev, float(fy26.RevUSDm), USD), (rOPM, float(fy26.OPM), PCT),
                         (rPAT, float(fy26.PATusdm), USD), (rNM, float(fy26.NetM), PCT),
                         (rFX, FX_FY26, '0.00'), (rRI, float(fy26.RevINRcr), INR),
                         (rPI, float(fy26.PATINRcr), INR), (rFCF, 42983.0, INR),
                         (rDIV, float(fy26.PayoutINRcr), INR), (rDPS, float(fy26.DPS), RUPEE),
                         (rTS, 28291.0, INR), (rRPE, round(REV_PER_EMP_FY26), '$#,##0'),
                         (rHC, int(fy26.Headcount), NUM)]:
        c = ws.cell(rr, 2, val); c.font = BLUE; c.number_format = fmt
    ws.cell(rEBIT, 2, f"=B{rRev}*B{rOPM}").number_format = USD

    for j in range(5):
        col = get_column_letter(3 + j)
        prev = get_column_letter(2 + j)
        acol = get_column_letter(2 + j)  # year column on Assumptions blocks
        set_ = [
            (rRev, f"='Revenue Build'!{col}{rb['total']}", USD),
            (rG, f"={col}{rRev}/{prev}{rRev}-1", PCT),
            (rOPM, f"={A}!{acol}{sr['marg']}", PCT),
            (rEBIT, f"={col}{rRev}*{col}{rOPM}", USD),
            (rPBT, f"={col}{rRev}*({col}{rOPM}+{A}!$B$10)", USD),
            (rPAT, f"={col}{rPBT}*(1-{A}!$B$11)", USD),
            (rNM, f"={col}{rPAT}/{col}{rRev}", PCT),
            (rFX, f"={prev}{rFX}*(1+{A}!$B$9)", '0.00'),
            (rRI, f"={col}{rRev}*{col}{rFX}/10", INR),
            (rPI, f"={col}{rPAT}*{col}{rFX}/10", INR),
            (rFCF, f"={col}{rPI}*{A}!{acol}{sr['fcf']}", INR),
            (rDIV, f"={col}{rFCF}*{A}!{acol}{sr['pay']}", INR),
            (rDPS, f"={col}{rDIV}/{A}!$B$13", RUPEE),
            (rTS, f"={col}{rDIV}*{A}!$B$12", INR),
            (rRPE, f"={prev}{rRPE}*(1+{A}!$B${sr['rpe']})", '$#,##0'),
            (rHC, f"={col}{rRev}*1000000/{col}{rRPE}", NUM),
            (rHCd, f"={col}{rHC}-$B${rHC}", NUM),
        ]
        for rr, f, fmt in set_:
            c = ws.cell(rr, 3 + j, f); c.number_format = fmt
            if rr in (rRev, rPAT, rDIV, rTS):
                c.font = BOLD

    rSum = rHCd + 2
    ws.cell(rSum, 1, "FY26–FY31 revenue CAGR").font = BOLD
    c = ws.cell(rSum, 3, f"=(G{rRev}/B{rRev})^(1/5)-1"); c.number_format = PCT; c.font = BOLD
    ws.cell(rSum + 1, 1, "Cumulative dividend to Tata Sons, FY27–FY31 (₹ crore)").font = BOLD
    c = ws.cell(rSum + 1, 3, f"=SUM(C{rTS}:G{rTS})"); c.number_format = INR; c.font = BOLD
    SO_ROWS[s] = dict(rev=rRev, pat=rPAT, div=rDIV, ts=rTS, hc=rHC, opm=rOPM,
                      cagr=rSum, cum=rSum + 1, dps=rDPS)
    r = rSum + 4

# probability-weighted block
header(ws, r, "Probability-weighted expected case", 7)
ws.cell(r + 1, 2, "FY26A").font = BOLD
ws.cell(r + 1, 2).fill = SUBFILL
for j, y in enumerate(YEARS):
    ws.cell(r + 1, 3 + j, y).font = BOLD
    ws.cell(r + 1, 3 + j).fill = SUBFILL
    ws.cell(r + 1, 3 + j).alignment = Alignment(horizontal="center")
wlines = [("Revenue ($mn)", "rev", USD), ("Net income ($mn)", "pat", USD),
          ("Total dividend (₹ crore)", "div", INR),
          ("Dividend to Tata Sons (₹ crore)", "ts", INR),
          ("Closing headcount", "hc", NUM)]
for k, (lab, key, fmt) in enumerate(wlines):
    rr = r + 2 + k
    ws.cell(rr, 1, lab).font = BOLD
    for j in range(6):
        col = get_column_letter(2 + j)
        terms = "+".join(f"{col}{SO_ROWS[s][key]}*{A}!$I${SC_ROWS[s]['prob'][0]}" for s in SCENS)
        c = ws.cell(rr, 2 + j, f"={terms}"); c.number_format = fmt; c.font = BOLD
rr = r + 2 + len(wlines) + 1
ws.cell(rr, 1, "Expected FY26–FY31 revenue CAGR").font = BOLD
c = ws.cell(rr, 3, f"=(G{r+2}/B{r+2})^(1/5)-1"); c.number_format = PCT; c.font = BOLD

# =============================================================================
# TATA SONS BRIDGE
# =============================================================================
ws = wb.create_sheet("Tata Sons Bridge")
ws.column_dimensions["A"].width = 52
for i in range(2, 9):
    ws.column_dimensions[get_column_letter(i)].width = 14
ws["A1"] = "Tata Sons — parent-level cash bridge (₹ crore)"
ws["A1"].font = TITLE
ws["A2"] = "Does the TCS dividend still fund the group's strategic bets? Capital deployment is held "\
           "constant across scenarios: the commitments do not shrink because TCS slows."
ws["A2"].font = NOTE

TSB = {}
r = 4
for s in SCENS:
    header(ws, r, s, 7)
    ws.cell(r + 1, 2, "FY26A").font = BOLD
    ws.cell(r + 1, 2).fill = SUBFILL
    for j, y in enumerate(YEARS):
        ws.cell(r + 1, 3 + j, y).font = BOLD
        ws.cell(r + 1, 3 + j).fill = SUBFILL
        ws.cell(r + 1, 3 + j).alignment = Alignment(horizontal="center")
    b = r + 2
    labs = ["Dividend from TCS", "Dividend from other listed holdings", "Total dividend inflow",
            "Parent opex, interest and tax", "Equity deployment into strategic bets",
            "Surplus / (deficit)", "Cumulative net cash", "TCS as % of dividend inflow",
            "Dividend cover of capital call (×)"]
    for k, lab in enumerate(labs):
        ws.cell(b + k, 1, lab).font = BOLD if k in (5, 6) else BLACK
    rTCS, rOTH, rTOT, rOPX, rCAP, rSUR, rCUM, rPCT, rCOV = [b + k for k in range(9)]
    # FY26 actual column
    for rr, v, fmt in [(rTCS, 28291, INR), (rOTH, TS_NON_TCS_DIV_FY26, INR),
                       (rOPX, TS_OPEX_TAX_FY26, INR), (rCAP, 15956, INR),
                       (rCUM, TS_NET_CASH_FY26, INR)]:
        c = ws.cell(rr, 2, float(v)); c.font = BLUE; c.number_format = fmt
    ws.cell(rTOT, 2, f"=B{rTCS}+B{rOTH}").number_format = INR
    ws.cell(rSUR, 2, f"=B{rTOT}-B{rOPX}-B{rCAP}").number_format = INR
    ws.cell(rPCT, 2, f"=B{rTCS}/B{rTOT}").number_format = PCT
    ws.cell(rCOV, 2, f"=B{rTOT}/B{rCAP}").number_format = '0.00"x"'
    for j in range(5):
        col = get_column_letter(3 + j)
        prev = get_column_letter(2 + j)
        acol = get_column_letter(2 + j)
        rows_ = [
            (rTCS, f"='Scenario Output'!{col}{SO_ROWS[s]['ts']}", INR),
            (rOTH, f"={prev}{rOTH}*(1+{A}!$B${TS0+1})", INR),
            (rTOT, f"={col}{rTCS}+{col}{rOTH}", INR),
            (rOPX, f"={prev}{rOPX}*(1+{A}!$B${TS0+3})", INR),
            (rCAP, f"={A}!{acol}{TS_CAP_ROW}", INR),
            (rSUR, f"={col}{rTOT}-{col}{rOPX}-{col}{rCAP}", INR),
            (rCUM, f"={prev}{rCUM}+{col}{rSUR}", INR),
            (rPCT, f"={col}{rTCS}/{col}{rTOT}", PCT),
            (rCOV, f"={col}{rTOT}/{col}{rCAP}", '0.00"x"'),
        ]
        for rr, f, fmt in rows_:
            c = ws.cell(rr, 3 + j, f); c.number_format = fmt
            if rr in (rSUR, rCUM):
                c.font = BOLD
    ws.cell(rCOV + 2, 1, "Cumulative surplus / (deficit) FY27–FY31").font = BOLD
    c = ws.cell(rCOV + 2, 3, f"=SUM(C{rSUR}:G{rSUR})"); c.number_format = INR; c.font = BOLD
    TSB[s] = dict(sur=rSUR, cum=rCUM, cov=rCOV, cumtot=rCOV + 2)
    r = rCOV + 5

# =============================================================================
# SENSITIVITY
# =============================================================================
ws = wb.create_sheet("Sensitivity")
ws.column_dimensions["A"].width = 40
for i in range(2, 10):
    ws.column_dimensions[get_column_letter(i)].width = 13
ws["A1"] = "Sensitivity — FY31 dividend to Tata Sons (₹ crore)"
ws["A1"].font = TITLE
ws["A2"] = "Rows: FY26–FY31 revenue CAGR (constant currency).  Columns: FY31 operating margin."
ws["A2"].font = NOTE

ws["A4"] = "Fixed inputs for this grid"
ws["A4"].font = BOLD
put(5, "FY26 revenue ($mn)", None); ws["B5"] = f"={A}!B5"; ws["B5"].font = GREEN; ws["B5"].number_format = USD
ws["A5"] = "FY26 revenue ($mn)"
ws["A6"] = "FY31 ₹/US$"
ws["B6"] = f"={A}!B8*(1+{A}!B9)^5"; ws["B6"].font = GREEN; ws["B6"].number_format = '0.00'
ws["A7"] = "Free cash flow / net income"
ws["B7"] = 0.82; ws["B7"].font = BLUE; ws["B7"].number_format = PCT; ws["B7"].fill = YELLOW
ws["A8"] = "Dividend as % of free cash flow"
ws["B8"] = 0.92; ws["B8"].font = BLUE; ws["B8"].number_format = PCT; ws["B8"].fill = YELLOW
for rr in range(5, 9):
    ws.cell(rr, 1).font = BLACK

G0 = 11
ws.cell(G0 - 1, 1, "FY31 dividend to Tata Sons (₹ crore)").font = BOLD
margins = [0.175, 0.190, 0.205, 0.220, 0.235, 0.250, 0.265]
cagrs = [-0.09, -0.06, -0.03, 0.00, 0.02, 0.04, 0.06, 0.08, 0.10]
ws.cell(G0, 1, "Revenue CAGR  \\  FY31 margin").font = BOLD
ws.cell(G0, 1).fill = SUBFILL
for j, m in enumerate(margins):
    c = ws.cell(G0, 2 + j, m); c.font = BLUE; c.number_format = PCT
    c.fill = SUBFILL; c.alignment = Alignment(horizontal="center")
for i, g in enumerate(cagrs):
    rr = G0 + 1 + i
    c = ws.cell(rr, 1, g); c.font = BLUE; c.number_format = PCT; c.fill = SUBFILL
    for j, m in enumerate(margins):
        col = get_column_letter(2 + j)
        f = (f"=$B$5*(1+$A{rr})^5*({col}${G0}+{A}!$B$10)*(1-{A}!$B$11)"
             f"*$B$6/10*$B$7*$B$8*{A}!$B$12")
        c = ws.cell(rr, 2 + j, f); c.number_format = INR; c.border = BOX
ws.cell(G0 + len(cagrs) + 2, 1,
        "Reference: Tata Sons received ₹28,291 crore from TCS in FY26 and ₹32,184 crore in FY25. "
        "Cells below ~₹28,000 crore mean the parent's dividend income does not recover to FY26 levels by FY31.").font = NOTE
ws.cell(G0 + len(cagrs) + 2, 1).alignment = Alignment(wrap_text=True, vertical="top")

# break-even
BE = G0 + len(cagrs) + 5
ws.cell(BE, 1, "Break-even check").font = BOLD
ws.cell(BE, 1).fill = SUBFILL
ws.cell(BE + 1, 1, "FY31 capital call on Tata Sons (₹ crore)").font = BLACK
ws.cell(BE + 1, 2, f"={A}!F{TS_CAP_ROW}").font = GREEN
ws.cell(BE + 1, 2).number_format = INR
ws.cell(BE + 2, 1, "Plus parent opex, interest and tax (₹ crore)").font = BLACK
ws.cell(BE + 2, 2, f"={A}!B{TS0+2}*(1+{A}!B{TS0+3})^5").font = GREEN
ws.cell(BE + 2, 2).number_format = INR
ws.cell(BE + 3, 1, "Less non-TCS dividend income (₹ crore)").font = BLACK
ws.cell(BE + 3, 2, f"=-{A}!B{TS0}*(1+{A}!B{TS0+1})^5").font = GREEN
ws.cell(BE + 3, 2).number_format = INR
ws.cell(BE + 4, 1, "TCS dividend to Tata Sons needed to break even in FY31").font = BOLD
ws.cell(BE + 4, 2, f"=SUM(B{BE+1}:B{BE+3})").font = BOLD
ws.cell(BE + 4, 2).number_format = INR
ws.cell(BE + 4, 2).fill = YELLOW

# =============================================================================
# RISK REGISTER
# =============================================================================
ws = wb.create_sheet("Risk Register")
for w, col in [(44, "A"), (13, "B"), (13, "C"), (13, "D"), (14, "E"), (14, "F"), (70, "G")]:
    ws.column_dimensions[col].width = w
ws["A1"] = "Revenue at risk by exposure bucket"
ws["A1"].font = TITLE
ws["A2"] = "Gross revenue at risk = FY26 revenue in the bucket × the share of its work AI can substantially "\
           "automate by FY31. It is a measure of exposure, not of expected loss: what is actually lost depends "\
           "on how much of the productivity gain is passed to the client and how far volume expands in response."
ws["A2"].font = NOTE
ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
ws.row_dimensions[2].height = 30

hdrs = ["Exposure bucket", "FY26 weight", "FY26 $mn", "Automatable share",
        "Gross rev at risk $mn", "Severe-case FY31 $mn", "Assessment"]
header(ws, 4, "", len(hdrs))
for j, h in enumerate(hdrs, 1):
    ws.cell(4, j, h)
    ws.cell(4, j).font = H1
    ws.cell(4, j).alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
ws.row_dimensions[4].height = 32
sev = RB_ROWS["Severe — Structural Break"]
for i, (name, w, auto, note) in enumerate(BUCKETS):
    r = 5 + i
    ws.cell(r, 1, name).font = BLACK
    ws.cell(r, 1).alignment = Alignment(wrap_text=True, vertical="top")
    ws.cell(r, 2, f"={A}!B{BR0+i}").number_format = PCT
    ws.cell(r, 3, f"={A}!C{BR0+i}").number_format = USD
    ws.cell(r, 4, f"={A}!D{BR0+i}").number_format = PCT
    ws.cell(r, 5, f"={A}!E{BR0+i}").number_format = USD
    ws.cell(r, 6, f"='Revenue Build'!G{sev['b0']+i}").number_format = USD
    ws.cell(r, 7, note).font = NOTE
    ws.cell(r, 7).alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[r].height = 46
rT2 = 5 + len(BUCKETS)
ws.cell(rT2, 1, "Total").font = BOLD
for col in (2, 3, 5, 6):
    c = ws.cell(rT2, col, f"=SUM({get_column_letter(col)}5:{get_column_letter(col)}{rT2-1})")
    c.font = BOLD
    c.number_format = PCT if col == 2 else USD
ws.cell(rT2 + 2, 1, "Gross revenue at risk as % of FY26 revenue").font = BOLD
ws.cell(rT2 + 2, 3, f"=E{rT2}/C{rT2}").number_format = PCT
ws.cell(rT2 + 2, 3).font = BOLD
ws.cell(rT2 + 2, 3).fill = YELLOW

for s in wb.worksheets:
    style_all(s)
    s.sheet_view.showGridLines = False
    s.freeze_panes = "B5" if s.title not in ("Read Me",) else None

wb.save("/root/tcs/TCS_Scenario_Model_FY27-FY31.xlsx")
print("written")
